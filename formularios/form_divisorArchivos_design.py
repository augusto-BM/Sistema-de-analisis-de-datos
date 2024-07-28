import os
import subprocess
import tkinter as tk
from tkinter import Tk, Label, Button, filedialog, Entry, messagebox, Frame, LabelFrame, Radiobutton
from openpyxl import load_workbook, Workbook  # Agregar Workbook aquí
import csv  # Importar el módulo csv
from config import COLOR_CUERPO_PRINCIPAL

class DivisorArchivosDesign():
    def __init__(self, panel_principal):
        self.panel_principal = panel_principal
        self.create_gui_divisorArchivos()

    def create_gui_divisorArchivos(self):
        self.barra_superior = tk.Frame(self.panel_principal)
        self.barra_superior.pack(side=tk.TOP, fill=tk.X, expand=False)

        self.labelTitulo = tk.Label(self.barra_superior, text="Divisor de Archivos Excel", padx=10, pady=10)
        self.labelTitulo.config(fg="#A7AB95", font=("Anton", 28), bg=COLOR_CUERPO_PRINCIPAL)
        self.labelTitulo.pack(side=tk.TOP, fill='both', expand=True)

        file_frame = LabelFrame(self.panel_principal, text="Selecciona un archivo de Excel", padx=10, pady=10)
        file_frame.pack(pady=5, padx=90, fill="both")

        file_button_frame = Frame(file_frame)
        file_button_frame.pack(pady=5, padx=5)

        self.choose_file_button = Button(file_button_frame, text="Seleccionar archivo", command=self.choose_file)
        self.choose_file_button.pack(pady=5, padx=10)

        self.file_path = ""
        self.file_path_label = Label(file_frame, text="", wraplength=800)
        self.file_path_label.pack(pady=5, padx=5, anchor="w")

        self.file_records_label = Label(file_frame, text="")
        self.file_records_label.pack(pady=5, padx=5, anchor="w")

        config_frame = LabelFrame(self.panel_principal, text="Configuración de División", padx=10, pady=10)
        config_frame.pack(pady=5, padx=90, fill="both")

        self.records_per_file_label = Label(config_frame, text="Registros por archivo:")
        self.records_per_file_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.records_per_file_entry = Entry(config_frame)
        self.records_per_file_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w", ipadx=50)
        self.records_per_file_entry.bind("<KeyRelease>", self.validate_number_input)

        folder_frame = LabelFrame(self.panel_principal, text="Carpeta de destino:", padx=10, pady=10)
        folder_frame.pack(pady=5, padx=90, fill="both")

        folder_button_frame = Frame(folder_frame)
        folder_button_frame.pack(pady=5, padx=5)

        self.choose_folder_button = Button(folder_button_frame, text="Seleccionar carpeta", command=self.choose_folder)
        self.choose_folder_button.pack(pady=5, padx=10)

        self.output_folder_path = ""
        self.output_folder_path_label = Label(folder_frame, text="", wraplength=500)
        self.output_folder_path_label.pack(pady=5, padx=5, anchor="w")

        format_frame = LabelFrame(self.panel_principal, text="Formato de Salida", padx=10, pady=10)
        format_frame.pack(pady=5, padx=90, fill="both")

        self.output_format = tk.StringVar(value="excel")

        excel_radio = Radiobutton(format_frame, text="Excel", variable=self.output_format, value="excel")
        excel_radio.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        csv_radio = Radiobutton(format_frame, text="CSV", variable=self.output_format, value="csv")
        csv_radio.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        self.divide_button = Button(self.panel_principal, text="Dividir archivo", command=self.divide_excel, width=20, bg="#A7AB95", fg="white", relief=tk.RAISED, cursor="hand2", pady=10)
        self.divide_button.pack(pady=5, padx=10)

    def choose_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            try:
                wb = load_workbook(filename=self.file_path, read_only=True)
                sheet = wb.active
                total_records = sheet.max_row - 1  # Restar 1 para excluir el encabezado
                self.file_path_label.config(text=f"Ruta de origen: {self.file_path}")
                self.file_records_label.config(text=f"Cantidad de registros: {total_records}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")

    def choose_folder(self):
        self.output_folder_path = filedialog.askdirectory()
        if self.output_folder_path:
            self.output_folder_path_label.config(text=self.output_folder_path)

    def validate_number_input(self, event):
        current_value = self.records_per_file_entry.get()
        if not current_value.isdigit() and current_value != "":
            self.records_per_file_entry.delete(0, 'end')
            messagebox.showwarning("Atención", "Por favor ingresa solo números.")

    def divide_excel(self):
        try:
            if not self.file_path:
                raise ValueError("Por favor selecciona un archivo de Excel.")

            records_per_file_entry_value = self.records_per_file_entry.get().strip()
            if not records_per_file_entry_value:
                raise ValueError("No puedes dejar vacio la cantidad de registros por archivo.")

            records_per_file = int(records_per_file_entry_value)
            if records_per_file <= 0:
                raise ValueError("La cantidad de registros por archivo debe ser mayor que cero.")

            wb = load_workbook(filename=self.file_path, read_only=True)
            sheet = wb.active
            total_records = sheet.max_row - 1  # Restar 1 para excluir el encabezado

            if records_per_file > total_records:
                raise ValueError(f"La cantidad de registros no puede ser mayor que {total_records}.")

            if not self.output_folder_path:
                raise ValueError("Selecciona una carpeta de destino para guardar los archivos.")

            num_files = (total_records + records_per_file - 1) // records_per_file

            base_filename, extension = os.path.splitext(os.path.basename(self.file_path))

            for i in range(num_files):
                start_row = i * records_per_file + 2  # +2 para saltar el encabezado
                end_row = min((i + 1) * records_per_file + 1, total_records + 1)

                if self.output_format.get() == "excel":
                    output_filename = f"{base_filename}({i + 1}).xlsx"
                    output_path = os.path.join(self.output_folder_path, output_filename)
                    new_wb = Workbook()
                    new_sheet = new_wb.active

                    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=sheet.max_column):
                        new_sheet.append([cell.value for cell in row])

                    new_wb.save(output_path)
                elif self.output_format.get() == "csv":
                    output_filename = f"{base_filename}({i + 1}).csv"
                    output_path = os.path.join(self.output_folder_path, output_filename)
                    
                    with open(output_path, 'w', newline='') as f:
                        writer = csv.writer(f)
                        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=sheet.max_column):
                            writer.writerow([cell.value for cell in row])

            messagebox.showinfo("Éxito", f"El archivo se ha dividido en {num_files} partes correctamente.")

            self.records_per_file_entry.delete(0, 'end')
            self.file_path_label.config(text="")
            self.file_records_label.config(text="")
            self.output_folder_path_label.config(text="")

            if os.name == 'nt':
                os.startfile(self.output_folder_path)
            elif os.name == 'posix':
                subprocess.Popen(['xdg-open', self.output_folder_path])

        except ValueError as ve:
            messagebox.showerror("Error de Valor", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")

